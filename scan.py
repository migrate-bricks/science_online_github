# -*- coding: utf-8 -*-
# before using this script, please refer to d = u2.connect("AUE66HL7XWIVJRSS") Change it based on 'adb devices'

import json
import logging
import os
import random
import re
import sys
import time
from datetime import datetime
import subprocess
from urllib.parse import urlparse
import colorlog
import openpyxl
import openpyxl.utils
import uiautomator2 as u2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import requests
from retrying import retry
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from ext.webdriver_manage_extend import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from uiautomator2.xpath import XMLElement

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

formatter = colorlog.ColoredFormatter(
    fmt='%(log_color)s %(asctime)s %(levelname)s:%(message)s%(reset)s',
    datefmt="%Y-%m-%d %H:%M:%S",
    log_colors={'DEBUG': 'cyan', 'INFO': 'green', 'WARNING': 'yellow', 'ERROR': 'red', 'CRITICAL': 'red,bg_white'}
)

console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

package_name = "com.taobao.idlefish"
browser_package_name = "com.android.browser"
activity_name = ".maincontainer.activity.MainActivity"


class TimeUtil:
    @staticmethod
    def random_sleep(random_start=2, random_end=5):
        wait_time = random.randint(random_start, random_end)
        time.sleep(wait_time)

    @staticmethod
    def sleep(secs: float) -> None:
        time.sleep(secs)

    @staticmethod
    def curr_date() -> str:
        return datetime.now().strftime("%Y-%m-%d")


def read_excel(excel_path: str, header_row: int) -> list:
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True, keep_links=False, rich_text=False)

    # Select the worksheet you want to read
    ws = wb.active

    # Get the column names
    columns = [cell.value for cell in next(ws.iter_rows(header_row))]  # second row is header

    # Read the data
    results = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if row[0] is None:
            break
        results.append({columns[i]: value for i, value in enumerate(row)})
    return results


def append_excel(data_list: list, output_file: str) -> None:
    if len(data_list) <= 0:
        return False

    if not os.path.exists(output_file):
        save_excel(data_list, output_file)
    else:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        ws.append(data_list.values())
        wb.save(output_file)


def save_excel(data_list: list, output_file: str) -> None:
    if data_list is None or len(data_list) <= 0:
        return False

    # Create file if not exist
    if not os.path.exists(os.path.dirname(output_file)):
        os.makedirs(os.path.dirname(output_file))

    # Open a new workbook
    wb = Workbook()
    ws = wb.active

    # Write the column header and apply formatting, set the width of each column based on the length of its name
    headers = list(data_list[0].keys())
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = PatternFill(start_color='00c5d9f1', end_color='00c5d9f1', fill_type='solid')
        ws.column_dimensions[cell.column_letter].width = max(10, len(header))
        # ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(10, len(header))

    # Write the data
    for row, row_data in enumerate(data_list, 2):
        for column, value in enumerate(row_data.values(), 1):
            ws.cell(row=row, column=column, value=value)

    # Determine the maximum number of columns
    max_column = max(len(row) for row in ws.rows)

    # Set the autofilter for all columns
    range_string = f"A1:{openpyxl.utils.get_column_letter(max_column)}{ws.max_row}"
    ws.auto_filter.ref = range_string
    # Freeze the top row
    ws.freeze_panes = ws['A2']
    wb.save(output_file)


def get_save_path(*paths: str) -> str:
    date = TimeUtil.curr_date()
    return os.path.join(os.getcwd(), 'save', date, *paths)


def swipe_up():
    d.swipe_ext("up", scale=0.9)


def open_page_by_keyword(search_keyword: str):
    d.app_start(package_name, activity_name, wait=True)
    d.xpath('//*[@resource-id="com.taobao.idlefish:id/search_bar_layout"]').wait()
    d.xpath('//*[@resource-id="com.taobao.idlefish:id/search_bar_layout"]').click()
    d.send_keys(search_keyword, clear=True)
    d.press('enter')


def open_page_by_url_old(url: str):  # Use browser to open store home page
    d.app_stop(browser_package_name)
    d.app_stop(package_name)
    d.app_start(browser_package_name, wait=True)
    d(resourceId="com.android.browser:id/search_hint").must_wait()
    d(resourceId="com.android.browser:id/search_hint").click_exists()
    d(resourceId="com.android.browser:id/url").must_wait()
    d(resourceId="com.android.browser:id/url").set_text(url)
    d.press('enter')
    d(textContains='允许').wait(exists=True, timeout=2)
    d(textContains='允许').click_exists()
    d(resourceId="android.miui:id/app1").wait(exists=True, timeout=2)
    d(resourceId="android.miui:id/app1").click_exists()
    d.sleep(5)


def open_page_by_url(url: str):  # Use adb shell to open store home page via app webview, performace is better
    intent_command = f'am start -n "{package_name}/{activity_name}" -d "{url}"'
    subprocess.run(["adb", "shell", intent_command], check=True)
    d(resourceId="android.miui:id/app1").wait(exists=True, timeout=2)  # In case there are more than one same app installed, it will popup the window
    d(resourceId="android.miui:id/app1").click_exists()
    d.sleep(1)


def get_platform_price(s: str) -> float:
    # Platform price format is different from Store price format
    match = re.search(r'¥(\d+\.?\d*)', s)
    if match:
        price = match.group(1)
        return float(price)


def get_store_price(s: str) -> float:
    # Platform price format is different from Store price format
    match = re.search(r'商品价格(\d+\.?\d*)', s)
    if match:
        price = match.group(1)
        return float(price)
    return 0


def get_wanted(s: str) -> float:
    match = re.search(r'(\d+\.?\d*)人想要', s)
    if match:
        price = match.group(1)
        return float(price)
    return 0


def get_min_price_but_greater_than_one(results: list) -> float:
    # Get the smallest price but skip those < 1, it's meaningless to do <1 biz
    min_price = sys.maxsize
    for item in results:
        if 1 <= item['price'] and item['price'] < min_price:
            min_price = item['price']
    return 0


def get_comebine_prices(results: list) -> str:
    sorted_results = sorted(results, key=lambda x: x['price'])
    prices = [str(item['price']) for item in sorted_results]
    return "_".join(sorted(prices))


def get_store_result_by_key(store_results: list, key: str):
    for store in store_results:
        if key.lower() in store['title'].lower():
            return store
    return {}


def clean_platform_text(text: str) -> str:
    return text.replace('\n', '')


def clean_text(text: str) -> str:
    return text.replace('\n', '@')


def main_complete():
    d.set_fastinput_ime(False)


def should_scan_store(store_excel_file_path: str) -> bool:
    if not os.path.exists(store_excel_file_path):
        return True
    store_name = os.path.splitext(os.path.basename(store_excel_file_path))[0]
    print(f'-> {store_name} results are already exists, path: {store_excel_file_path}')
    overwrite = input('【Overwrite? 1.Yes, 2.No: ')
    return (overwrite == '1')


def scan_platform(idx: int, search_keywords: str, must_include_word: str, max_scroll_page: int, scroll_page_timeout: float):
    try:
        logger.info(f"Retrieving products information for 【{search_keywords}】")
        results = []
        for search_keyword in search_keywords:
            open_page_by_keyword(search_keyword)
            for i in range(max_scroll_page):
                logger.info(f"Scrolling to idx: {idx}, keyword: {search_keyword}, include: {must_include_word} [{i}/{max_scroll_page}] page")
                TimeUtil.sleep(scroll_page_timeout)
                view_list = d.xpath('//android.widget.ScrollView//android.view.View').all()
                if len(view_list) > 0:
                    for el in view_list:
                        if len(el.elem.getchildren()) > 0:
                            el_description = clean_platform_text(str(el.attrib['content-desc']))
                            for child in el.elem.getchildren():
                                el_description = f"{el_description}&{clean_platform_text(str(child.attrib['content-desc']))}"  # combine el_description
                            if must_include_word.lower() in el_description.lower():
                                price = get_platform_price(el_description)
                                wanted = get_wanted(el_description)
                                # skip duplicated item
                                if price is not None and price != '' and not any(d['title'] == el_description for d in results):
                                    logger.info(f"【{len(results)+1}】-description:{el_description}, price:{price}, wanted:{wanted}")
                                    results.append({'title': el_description, 'price': price, 'wanted': wanted})
                if d(descriptionContains='到底了').exists:  # alread on the end of the page
                    break
                swipe_up()
        return results
    except Exception as e:
        print(e)
        logger.error("Program runs Error:" + str(e.args[0]))
    finally:
        main_complete()
        print("Execution Completed!")


def retrieve_store_detail(eleView: XMLElement) -> str | None:
    eleView.click()
    d.xpath("//*[@content-desc='分享']").wait()
    d.xpath("//*[@content-desc='分享']").click()
    d.xpath("//*[@content-desc='复制链接']").wait()
    d.xpath("//*[@content-desc='复制链接']").click()
    url = extract_url_from_text(d.xpath("//*[starts-with(@content-desc, '【闲鱼】')]").attrib['content-desc'])
    d.press("back")
    d.xpath("//*[starts-with(@content-desc, '【闲鱼】')]").wait_gone()
    d.press("back")
    d.xpath('//*[@content-desc="管理"]').wait_gone()
    return parse_html_page(url)


def extract_url_from_text(text):
    url_pattern = re.compile(
        r'http[s]?://'
        r'(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
    )
    match = url_pattern.search(text)
    return match.group(0) if match else None


def scan_store(store_name: str, must_include_word: str, max_scroll_page: int, scroll_page_timeout: float):
    try:
        logger.info(f"【Retrieving products information for {store_name}")
        store_results = []
        for idx in range(max_scroll_page):
            logger.info(f"【Scrolling to store: {store_name} [{idx}/{max_scroll_page}] page")
            TimeUtil.sleep(scroll_page_timeout)
            view_list = d.xpath('//android.widget.ScrollView//android.view.View').all()
            if len(view_list) > 0:
                for el in view_list:
                    title = clean_text(str(el.attrib['content-desc']))
                    if must_include_word.lower() in title.lower():
                        # {"product_url": product_url, "soldprice": soldprice, "wants": wants, "likes": likes, "views": views}
                        details = retrieve_store_detail(el)
                        product_url = details["product_url"]
                        soldprice = details["soldprice"]
                        wants = details["wants"]
                        likes = details["likes"]
                        views = details["views"]
                        if details and all(d['title'] != title for d in store_results):  # Skip duplicated item
                            product = {"title": title, "product_url": product_url, "soldprice": soldprice, "wants": wants, "likes": likes, "views": views}
                            store_results.append(product)
                            logger.info(f"【{len(store_results)+1}】- {product}")

            if d(descriptionContains='没有更多了').exists:
                break
            swipe_up()
        return store_results
    except Exception as e:
        print(e)
        logger.error("Program runs Error:" + str(e.args[0]))
    finally:
        main_complete()
        print("Execution Completed!")


def load_store_results(store_name: str, store_homepage: str, store_must_include_word: str, store_max_scroll_page: int, store_scroll_page_timeout: float) -> list:
    store_excel_file_path = get_save_path(f"STORE_{store_name}.xlsx")
    if should_scan_store(store_excel_file_path):
        print(f'【Navigate to store: {store_name}, homepage: {store_homepage}')
        open_page_by_url(store_homepage)
        print(f'【Scan store: {store_name}, homepage: {store_homepage}')
        store_results = scan_store(store_name, store_must_include_word, store_max_scroll_page, store_scroll_page_timeout)
        store_sorted_results = sorted(store_results, key=lambda x: x['wanted'], reverse=True)
        print(f'【Save store {store_name} results, path: {store_excel_file_path}')
        save_excel(store_sorted_results, store_excel_file_path)
    else:
        print(f'【Read store {store_name} results from existing path: {store_excel_file_path}')
        store_results = read_excel(store_excel_file_path, header_row=1)
    return store_results


def full_outer_join(store_results: list, delivery_settings: list):
    for store in store_results:  # Set must_include_word on store results
        for delivery in delivery_settings:
            if delivery['must_include_word'].lower() in store['title'].lower():
                store['must_include_word'] = delivery['must_include_word']
                break
    resultsDataFrame = pd.merge(pd.DataFrame(delivery_settings), pd.DataFrame(store_results), on='must_include_word', how='outer')
    return resultsDataFrame.to_dict(orient='records')


@retry(stop_max_attempt_number=3, wait_fixed=1500)
def fetch_url_retry(url, stream):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
    response = requests.Session().get(url=url, headers=headers, stream=stream)
    response.raise_for_status()
    return response


def smooth_scroll_to(driver, element, duration=1500):
    js_code = """
        function smoothScrollTo(element, duration) {
            const startingY = window.pageYOffset;
            const elementY = element.getBoundingClientRect().top + window.pageYOffset;
            const targetY = document.body.scrollHeight - elementY < window.innerHeight ? document.body.scrollHeight - window.innerHeight : elementY;
            const diff = targetY - startingY;
            let start;

            function step(timestamp) {
                if (!start) start = timestamp;
                const time = timestamp - start;
                const percent = Math.min(time / duration, 1);
                window.scrollTo(0, startingY + diff * percent);
                if (time < duration) {
                    window.requestAnimationFrame(step);
                }
            }

            window.requestAnimationFrame(step);
        }
        smoothScrollTo(arguments[0], arguments[1]);
    """
    driver.execute_script(js_code, element, duration)
    time.sleep(duration/1000)  # because the above code is async on ui render, need to wait the duration time


def sanitize_filename(filename):
    invalid_chars_pattern = r'[\\/:*?"<>|\r\n\t]'
    sanitized_filename = re.sub(invalid_chars_pattern, '_', filename)
    sanitized_filename = sanitized_filename.strip('. ')
    return sanitized_filename


def parse_html_page(url):
    service = Service(ChromeDriverManager().install())
    chrome_options = Options()
    # chrome_options.add_argument('--headless')
    # chrome_options.add_argument('--disable-gpu')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument("--disable-blink-features")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        # Escape Taobao's anti-crawling mechanism
        script = '''
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            })
        '''
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": script})

        driver.get(url)

        xpath_qrcode_container = "//div[starts-with(@class, 'rax-view-v2 Detail--qrcodeContainer--')]"
        qrcode_container = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_qrcode_container)))
        driver.execute_script("arguments[0].remove();", qrcode_container)

        # It's iframe website, support scrolling need to go to iframe first
        iframe_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "iframe")))
        driver.switch_to.frame(iframe_element)

        xpath_detail_desc_expression = "//span[starts-with(@class, 'rax-text-v2 detailDesc--descText--')]"
        detail_desc_span = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_detail_desc_expression)))

        sub_folder = get_save_path(sanitize_filename(detail_desc_span.text[:30])).replace('\n', ' ')
        if not os.path.exists(sub_folder):
            os.makedirs(sub_folder)

        detail_desc_filepath = get_save_path(sub_folder, "doc.txt")
        with open(detail_desc_filepath, 'w', encoding='utf-8') as file:
            file.write(detail_desc_span.text)

        max_scroll_attempts = 20
        current_scroll_attempt = 0
        xpath_recommendwrap_expression = "//div[starts-with(@class, 'rax-view-v2 recommendGoods--recommendWrap--')]"

        while current_scroll_attempt < max_scroll_attempts:
            try:
                recommend_wrap = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_recommendwrap_expression)))
                smooth_scroll_to(driver, recommend_wrap)
                driver.execute_script("arguments[0].remove();", recommend_wrap)
                print(f"Found target element {recommend_wrap}")
                break
            except Exception:
                driver.execute_script("window.scrollBy(0, 300);")
                current_scroll_attempt += 1
                print(f"{current_scroll_attempt} attempt at scrolling...")

            if current_scroll_attempt == max_scroll_attempts:
                print("The maximum number of scrolls has been reached and the target element has not been found.")
                break

        mhtml_content = driver.execute_cdp_cmd("Page.captureSnapshot", {"format": "mhtml"})
        mhtml_data = mhtml_content['data'].encode('utf-8')
        mhtml_path = get_save_path(sub_folder, "index.mhtml")
        with open(mhtml_path, 'wb') as file:
            file.write(mhtml_data)
        print(f"mhtml page is saved to :{mhtml_path}")

        xpath_soldprice_expression = "//div[starts-with(@class, 'rax-text-v2 priceMod--soldPrice--')]"
        span_price = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_soldprice_expression)))
        if span_price:
            soldprice = int(span_price.get_attribute("textContent"))

        xpath_wantdetail_expression = "//div[starts-with(@class, 'rax-text-v2 subDetailMod--wantDetail--')]"
        span_wantdetail = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_wantdetail_expression)))
        if span_wantdetail:
            wantdetail = span_wantdetail.get_attribute("textContent")
            matches = re.findall(r"(\d+)人想要|赞(\d+)|浏览 (\d+)", wantdetail)
            wants = int(matches[0][0]) if matches[0] else None
            likes = int(matches[1][0]) if matches[1] else None
            views = int(matches[2][0]) if matches[2] else None

        product_url = get_base_url(driver.current_url)

        xpath_imagecontainer_expression = "//div[starts-with(@class, 'rax-view-v2 imageListMod--imageWrap--')]"
        image_container = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_imagecontainer_expression)))

        if image_container:
            image_elements = image_container.find_elements(By.TAG_NAME, "img")
            for index, image_element in enumerate(image_elements, start=1):
                image_url = image_element.get_attribute("src")
                if image_url and image_url.startswith("http"):
                    try:
                        image_name = f"{index}.png"
                        image_path = get_save_path(sub_folder, image_name)
                        response = fetch_url_retry(image_url, stream=True)
                        if response.status_code == 200:
                            with open(image_path, 'wb') as file:
                                for chunk in response.iter_content(chunk_size=1024):
                                    if chunk:
                                        file.write(chunk)
                            print(f"Picture is downloaded to: {image_path}")
                        else:
                            print(f"Picture download failed, URL: {image_url}, status: {response.status_code}")
                    except Exception as e:
                        print(f"Picture download exception: {e}")

        driver.switch_to.default_content()
    finally:
        driver.quit()
        return {"product_url": product_url, "soldprice": soldprice, "wants": wants, "likes": likes, "views": views}


def get_base_url(full_url):
    parsed_url = urlparse(full_url)
    base_url = f"{parsed_url.scheme}://{parsed_url.netloc}{parsed_url.path}"
    return base_url


def get_connected_devices():
    output = subprocess.check_output(['adb', 'devices']).decode('utf-8')
    devices_lines = output.strip().split('\n')[1:]
    devices = [line.split('\t')[0] for line in devices_lines if line]
    return devices


def get_device_details(device_sn):
    model_cmd = f'adb -s {device_sn} shell getprop ro.product.model'
    manufacturer_cmd = f'adb -s {device_sn} shell getprop ro.product.manufacturer'
    android_version_cmd = f'adb -s {device_sn} shell getprop ro.build.version.release'

    model = subprocess.check_output(model_cmd, shell=True).decode('utf-8').strip()
    manufacturer = subprocess.check_output(manufacturer_cmd, shell=True).decode('utf-8').strip()
    android_version = subprocess.check_output(android_version_cmd, shell=True).decode('utf-8').strip()

    return {
        "Model": model,
        "Manufacturer": manufacturer,
        "AndroidVersion": android_version
    }


if __name__ == '__main__':
    print('Please be sure: \n1.The uiautomator2 has connected to android device\n2.Setup the `correct android_device_addr` in scan_config.json\n')

    with open('./scan_config.json', 'r', encoding='utf8') as fp:
        scan_config = json.load(fp)

        scroll_page_timeout_second = scan_config['scroll_page_timeout_second']
        delivery_settings_path = scan_config['delivery_settings_path']

        platform_max_scroll_page = scan_config['scan_platform']['max_scroll_page']

        store_max_scroll_page = scan_config['scan_store']['max_scroll_page']
        store_must_include_word = scan_config['scan_store']['must_include_word']
        stores = scan_config['scan_store']['stores']

        # Define first store is the main store
        main_store_name = stores[0]['store_name']
        main_store_homepage = stores[0]['home_page']

        connected_devices = get_connected_devices()
        if connected_devices:
            d = u2.connect(connected_devices[0])  # Get first device id
        else:
            print("Not found devices...")
            exit

        scan_type = input('【Select scan type 1.Platform or 2.Store: ')

        if scan_type == '1':  # Scan platform
            print('【Scan platform Start')
            print(f'【Read global delivery settings from path: {delivery_settings_path}')
            delivery_settings = read_excel(delivery_settings_path, header_row=2)  # Header:*配置名称|*外部编码|*附言（具体的发货内容填写在这里）|商品分类（选填）|自动发货开关（不填默认开启）|配置名称是否等于外部编码|附言是否包含外部编码|search_keywords|must_include_word|

            # Retrieve Main store results
            print('【Load main store results for getting current price')
            main_store_results = load_store_results(main_store_name, main_store_homepage, store_must_include_word, store_max_scroll_page, scroll_page_timeout_second)

            summary_results = []
            for idx, delivery_setting in enumerate(delivery_settings):  # Excel settings is delivery_setting
                setting_search_keywords = delivery_setting['search_keywords'].split(',')
                setting_must_include_word = delivery_setting['must_include_word']  # Foreign key

                print(f'【Scan Platform by keyword: {setting_search_keywords}, include: {setting_must_include_word}')
                platform_results = scan_platform(idx, setting_search_keywords, setting_must_include_word, platform_max_scroll_page, scroll_page_timeout_second)

                if platform_results is not None:
                    platform_sorted_results = sorted(platform_results, key=lambda x: x['price'])
                    platform_combine_prices = get_comebine_prices(platform_results)
                    platform_min_price = get_min_price_but_greater_than_one(platform_results)

                # Find target main store result, get current title, wanted, price
                main_store_result = get_store_result_by_key(main_store_results, setting_must_include_word)
                if main_store_result is not None:
                    current_price = main_store_result.get('price')

                platform_excel_file_name = f"PLATFORM_{setting_must_include_word}_price_{current_price}_min_{platform_min_price}.xlsx"
                platform_excel_file_path = get_save_path(platform_excel_file_name)
                print(f'【{idx}_Save Platform results keywords:{setting_search_keywords}, include: {setting_must_include_word} path: {platform_excel_file_path}')
                save_excel(platform_sorted_results, platform_excel_file_path)
                logger.info(f"【{idx}: Save Platform results completed, File path: {platform_excel_file_path}")

                summary_result = {**delivery_setting, **main_store_result, 'min_price': platform_min_price, 'combine_prices': platform_combine_prices}
                summary_results.append(summary_result)
                logger.info(f'【Summary result: {summary_result}')
                if (len(summary_results) % 30 == 0):
                    save_excel(summary_results, get_save_path(f"SUMMARY_PLATFORM_{len(summary_results)}.xlsx"))

            summary_excel_file_path = get_save_path("SUMMARY_PLATFORM.xlsx")
            save_excel(summary_results, summary_excel_file_path)
            logger.info(f"【Platform Summary Execution completed, File path: {summary_excel_file_path}")
        elif scan_type == '2':  # Scan Store
            print('【ALL available stores:')
            for idx, store in enumerate(stores, start=1):
                print(f"-> {idx}.{store['store_name']} {store['home_page']}")
            store_index = input('【Select store index: ')
            store_name = stores[int(store_index)-1]['store_name']
            store_homepage = stores[int(store_index)-1]['home_page']

            print(f'【Read global delivery settings from path: {delivery_settings_path}')
            delivery_settings = read_excel(delivery_settings_path, header_row=2)  # Header:*配置名称|*外部编码|*附言（具体的发货内容填写在这里）|商品分类（选填）|自动发货开关（不填默认开启）|配置名称是否等于外部编码|附言是否包含外部编码|search_keywords|must_include_word|

            print('【Load Store results')
            store_results = load_store_results(store_name, store_homepage, store_must_include_word, store_max_scroll_page, scroll_page_timeout_second)

            merge_results = full_outer_join(store_results, delivery_settings)

            summary_store_excel_file_path = get_save_path(f"SUMMARY_STORE_{store_name}.xlsx")
            print(f'【Save {store_name} STORE SUMMARY results, path: {summary_store_excel_file_path}')
            save_excel(merge_results, summary_store_excel_file_path)
