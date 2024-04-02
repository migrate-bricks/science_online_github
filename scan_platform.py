# -*- coding: utf-8 -*-
# before using this script, please refer to d = u2.connect("AUE66HL7XWIVJRSS") Change it based on 'adb devices'

import json
import logging
from operator import itemgetter
import os
import pathlib
import random
import re
import sys
import time
from datetime import datetime, date, timedelta

import colorlog
import openpyxl
import uiautomator2 as u2
from openpyxl import Workbook

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

formatter = colorlog.ColoredFormatter(
    fmt='%(log_color)s %(asctime)s %(levelname)s:%(message)s%(reset)s',
    datefmt="%Y-%m-%d %H:%M:%S",
    log_colors={
        'DEBUG': 'cyan',
        'INFO': 'green',
        'WARNING': 'yellow',
        'ERROR': 'red',
        'CRITICAL': 'red,bg_white',
    }
)

console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# d = u2.connect("AUE66HL7XWIVJRSS")  # TODO: Change it based on 'adb devices'

package_name = "com.taobao.idlefish"
activity_name = ".maincontainer.activity.MainActivity"


class TimeUtil:
    @staticmethod
    def random_sleep(random_start=2, random_end=5):
        wait_time = random.randint(random_start, random_end)
        time.sleep(wait_time)

    @staticmethod
    def sleep(secs):
        time.sleep(secs)

    @staticmethod
    def curr_date():
        return datetime.now().strftime("%Y-%m-%d")

    @staticmethod
    def tomorrow_date():
        today = date.today()
        tomorrow = today + timedelta(days=1)
        return tomorrow.strftime('%Y-%m-%d')


# def get_desktop_path():
#     if sys.platform == 'win32':
#         desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
#     elif sys.platform == 'darwin':
#         desktop_path = str(pathlib.Path.home() / "Desktop")
#     else:
#         desktop_path = None
#     return desktop_path


# def excel_cell_to_index(cell_str):
#     """
#     Excel cells are converted to row and column indexes, for example, A1 is converted to 1,1 and C4 is converted to 3,4
#     :param cell_str:
#     :return:
#     """
#     letter = cell_str[0]
#     number = int(cell_str[1:])
#     column = ord(letter.upper()) - ord('A') + 1
#     row = number
#     return row, column


# def read_delivery_settings(excel_path):
#     wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True, keep_links=False, rich_text=False)
#     sh = wb['Sheet1']

#     columns = [cell.value for cell in next(sh.iter_rows(2))]  # second row is header
#     code_index = columns.index('*外部编码') + 1
#     search_keywords_index = columns.index('search_keywords') + 1
#     must_include_word_index = columns.index('must_include_word') + 1
#     results = []
#     for r in range(3, sh.max_row + 1):
#         code = sh.cell(row=r, column=code_index).value
#         search_keywords = sh.cell(row=r, column=search_keywords_index).value.split(',')
#         must_include_word = sh.cell(row=r, column=must_include_word_index).value
#         if code is None:
#             break
#         results.append({'code': code, 'search_keywords': search_keywords, 'must_include_word': must_include_word})
#     return results


def read_excel(excel_path):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path, read_only=True, keep_vba=False, data_only=True, keep_links=False, rich_text=False)

    # Select the worksheet you want to read
    sh = wb['Sheet1']

    # Get the column names
    columns = [cell.value for cell in next(sh.iter_rows(2))]  # second row is header

    # Read the data
    results = []
    for row in sh.iter_rows(min_row=3, values_only=True):
        results.append({columns[i]: value for i, value in enumerate(row)})
    return results


def save_excel(data_list, output_file):
    # Open a new workbook
    wb = Workbook()
    sh = wb.active

    # Write the column header
    headers = list(data_list[0].keys)
    for i, header in enumerate(headers, 1):
        sh.cell(row=1, column=i, value=header)

    # Write the data
    for row, row_data in enumerate(data_list, 2):
        for column, value in enumerate(row_data.values(), 1):
            sh.cell(row=row, column=column, value=value)
    wb.save(output_file)


def get_save_folder():
    date = TimeUtil.curr_date()
    return os.path.join(os.getcwd(), 'save', date)


def get_save_path(filename):
    return os.path.join(get_save_folder(), filename)


# def to_excel(data_list, output_file):
#     # dt = TimeUtil.curr_date()
#     # write_path = os.path.join(os.getcwd(), 'save')
#     # if not os.path.exists(write_path):
#     # os.makedirs(write_path)
#     wb = Workbook()
#     sheet = wb.active
#     sheet.column_dimensions["A"].width = 100
#     sheet_name = 'Sheet1'
#     sheet.title = sheet_name
#     sheet['A1'] = 'Title'
#     sheet['B1'] = 'Price'
#     sheet['C1'] = 'Wanted'
#     sheet['D1'] = 'Profit'
#     start_row = 2
#     sorted_data_list = sorted(data_list, key=itemgetter('price'), reverse=False)
#     # mindata = min(data_list, key=itemgetter('price'))
#     # output_file = os.path.join(write_path, f"{dt}-{keyword}-{mindata['price']}.xlsx")
#     for index, data in enumerate(sorted_data_list):
#         sheet["A" + str(index + start_row)] = data['title']
#         sheet["B" + str(index + start_row)] = data['price']
#         sheet["C" + str(index + start_row)] = data['wanted']
#         sheet["D" + str(index + start_row)] = data['price'] * data['wanted']
#     wb.save(filename=output_file)
#     return output_file


def swipe_up():
    d.swipe_ext('up', 0.9)


def open_page_by_keyword(search_keyword):
    d(resourceId="com.taobao.idlefish:id/title").must_wait()
    d(resourceId="com.taobao.idlefish:id/title").click()
    d.send_keys(search_keyword, clear=True)
    d.press('enter')


def get_price(s):
    match = re.search(r'¥(\d+\.?\d*)', s)
    if match:
        price = match.group(1)
        return float(price)


def get_wanted(s):
    match = re.search(r'(\d+\.?\d*)人想要', s)
    if match:
        price = match.group(1)
        return float(price)
    return 0


def get_min_price_but_greater_than_one(results):
    # Get the smallest price but skip those < 1, it's meaningless to do <1 biz
    min_price = sys.maxint
    for item in results:
        if 1 <= item['price'] and item['price'] < min_price:
            min_price = item['price']
    return min_price


def clean_text(text):
    return text.replace('\n', '')


def main_complete():
    d.set_fastinput_ime(False)


def execute_scan(search_keywords, must_include_word, max_scroll_page):
    try:
        logger.info(d.info)
        d.app_start(package_name, activity_name, wait=True)
        logger.info(f"Retrieving products information for 【{search_keywords}】...")
        results = []
        for search_keyword in search_keywords:
            open_page_by_keyword(search_keyword)
            for i in range(max_scroll_page):
                logger.info(f"Scrolling to [{i}/{max_scroll_page}] page...")
                TimeUtil.random_sleep()
                view_list = d.xpath('//android.widget.ScrollView//android.view.View').all()
                if len(view_list) > 0:
                    for el in view_list:
                        if len(el.elem.getchildren()) > 0:
                            el_description = clean_text(str(el.attrib['content-desc']))
                            for child in el.elem.getchildren():
                                el_description = f"{el_description}&{clean_text(str(child.attrib['content-desc']))}"  # combine el_description
                            if must_include_word.lower() in el_description.lower():
                                price = get_price(el_description)
                                wanted = get_wanted(el_description)
                                # skip duplicated item
                                if price is not None and price != '' and not any(d['title'] == el_description for d in results):
                                    logger.info(f"【{len(results)+1}】-description:{el_description}, price:{price}, wanted:{wanted}")
                                    results.append({'title': el_description, 'price': price, 'wanted': wanted})
                if d(descriptionContains='到底了').exists:  # alread on the end of the page
                    break
                swipe_up()
        # output_file = to_excel(results, must_include_word)
        # logger.info(f"Execution completed, file path: {output_file}")
        return results
    except Exception as e:
        print(e)
        logger.error("Program runs Error:" + str(e.args[0]))
    finally:
        main_complete()
        print("Execution Completed!")


if __name__ == '__main__':
    with open('./scan_platform_config.json', 'r', encoding='utf8') as fp:
        platform_config = json.load(fp)

        android_device_addr = platform_config["android_device_addr"]
        max_scroll_page = platform_config['max_scroll_page']
        delivery_settings_path = platform_config['delivery_settings_path']
        # searchs = read_delivery_settings(delivery_settings_path)
        searchs = read_excel(delivery_settings_path)

        d = u2.connect(android_device_addr)  # TODO: Change it based on 'adb devices' "AUE66HL7XWIVJRSS"

        for search in searchs:
            search_keywords = search['search_keywords']
            must_include_word = search['must_include_word']
            results = execute_scan(search_keywords=search_keywords, must_include_word=must_include_word, max_scroll_page=max_scroll_page)

            sorted_results = sorted(results, key=lambda x: x['price'])
            min_price = get_min_price_but_greater_than_one(results)
            excel_file_name = f"{must_include_word}-{min_price}.xlsx"
            save_excel(sorted_results, get_save_path(excel_file_name))
            # output_file = to_excel(results, must_include_word)
            logger.info(f"Execution completed, file path: {excel_file_name}")
